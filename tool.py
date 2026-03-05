import pandas as pd
from pathlib import Path
from tkinter import Tk, filedialog, messagebox
from datetime import datetime
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

warnings.filterwarnings("ignore")

# 🎨 Highlight colors

BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

def choose_file():
    Tk().withdraw()
    return filedialog.askopenfilename(
        title="Select Excel or CSV file",
        filetypes=[
            ("Excel Files", "*.xlsx *.xls"),
            ("CSV Files", "*.csv"),
            ("All Files", "*.*")
        ]
    )


def load_file(filepath):
    ext = Path(filepath).suffix.lower()

    if ext == ".csv":
        df = pd.read_csv(filepath)
    else:
        df = pd.read_excel(filepath)

    df.columns = df.columns.str.strip()
    return df


def find_class_column(df):
    possible_names = [
        "class",
        "ci class",
        "configuration item class",
        "asset class",
        "ci_class",
        "asset_class"
    ]

    normalized_columns = {
        col.lower().replace("_", " ").strip(): col
        for col in df.columns
    }

    for name in possible_names:
        if name in normalized_columns:
            return normalized_columns[name]

    raise Exception(
        "No Class column found.\n"
        "Ensure your file contains:\n"
        "Class / CI Class / Asset Class"
    )


def get_downloads_path():
    return Path.home() / "Downloads"


def highlight_sheet(ws):
    """Apply highlighting to missing values & duplicate rows"""

    # Convert worksheet to list for duplicate detection
    data = list(ws.values)
    headers = data[0]
    rows = data[1:]

    # detect duplicate rows
    duplicate_rows = set()
    seen = set()

    for idx, row in enumerate(rows, start=2):
        if row in seen:
            duplicate_rows.add(idx)
        else:
            seen.add(row)

    for row in ws.iter_rows(min_row=2):
        row_idx = row[0].row

        # highlight duplicate rows
        if row_idx in duplicate_rows:
            for cell in row:
                cell.fill = YELLOW_FILL

        # highlight missing cells
        for cell in row:
            if cell.value is None or str(cell.value).strip() == "":
                cell.fill = BLUE_FILL


def audit_data(df, class_col):

    downloads_folder = get_downloads_path()
    today_date = datetime.now().strftime("%Y-%m-%d")
    output_path = downloads_folder / f"CMDB_Audit_{today_date}.xlsx"

    writer = pd.ExcelWriter(output_path, engine="openpyxl")
    summary_data = []

    # Original data sheet
    df.to_excel(writer, sheet_name="Original_Data", index=False)

    for cls, group in df.groupby(class_col):
        group = group.copy()

        group["Duplicate"] = group.duplicated(keep=False)
        duplicate_count = group["Duplicate"].sum()

        duplicate_values = ""

        if duplicate_count > 0:
            preferred_cols = ["Name", "CI Name", "Hostname", "Asset Name"]

            found_col = None
            for col in preferred_cols:
                if col in group.columns:
                    found_col = col
                    break

            if found_col:
                dup_vals = group[group["Duplicate"]][found_col].dropna().astype(str).unique()
                duplicate_values = ", ".join(dup_vals[:10])
            else:
                duplicate_values = "Duplicate rows detected"

        missing_counts = {}
        for col in group.columns:
            missing = group[col].isna() | (group[col].astype(str).str.strip() == "")
            count = missing.sum()
            if count > 0:
                missing_counts[col] = count

        incomplete_rows = group.isna().any(axis=1).sum()

        missing_text = ", ".join(
            [f"{col} ({count})" for col, count in missing_counts.items()]
        ) if missing_counts else ""

        summary_data.append([
            cls,
            len(group),
            duplicate_count,
            duplicate_values,
            incomplete_rows,
            missing_text
        ])

        sheet_name = str(cls)[:31]
        group.to_excel(writer, sheet_name=sheet_name, index=False)

    summary_df = pd.DataFrame(
        summary_data,
        columns=[
            "Class",
            "Total Rows",
            "Duplicates",
            "Duplicate Values",
            "Incomplete Rows",
            "Columns with Missing Values"
        ]
    )

    summary_df.to_excel(writer, sheet_name="Audit_Summary", index=False)
    writer.close()

    # Apply highlighting & reorder sheets
    wb = load_workbook(output_path)

    for sheet in wb.sheetnames:
        if sheet not in ["Audit_Summary"]:
            highlight_sheet(wb[sheet])

    new_order = ["Audit_Summary", "Original_Data"] + [
        s for s in wb.sheetnames if s not in ["Audit_Summary", "Original_Data"]
    ]

    wb._sheets = [wb[s] for s in new_order]
    wb.save(output_path)

    messagebox.showinfo(
        "Audit Complete",
        f"Audit saved to Downloads:\n{output_path}"
    )


if __name__ == "__main__":
    file_path = choose_file()

    if not file_path:
        exit()

    df = load_file(file_path)
    class_column = find_class_column(df)
    audit_data(df, class_column)
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from pathlib import Path

# 🎨 Highlight colors
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
RED_FILL = PatternFill(start_color="F48FB1", end_color="F48FB1", fill_type="solid")

class ReportGenerator:
    def __init__(self, auditor, original_df):
        self.auditor = auditor
        self.original_df = original_df

    def generate_excel(self, output_path):
        writer = pd.ExcelWriter(output_path, engine="openpyxl")
        
        # 1. Summary Sheet
        summary_df = self.auditor.get_summary_df()
        summary_df.to_excel(writer, sheet_name="Audit_Summary", index=False)
        
        # 2. Original Data Sheet
        self.original_df.to_excel(writer, sheet_name="Original_Data", index=False)
        
        # 3. Group Sheets
        for cls, group in self.original_df.groupby(self.auditor.class_col) if self.auditor.class_col else [("Data", self.original_df)]:
            sheet_name = str(cls)[:31]
            group.to_excel(writer, sheet_name=sheet_name, index=False)
        
        writer.close()
        
        # 4. Post-process Excel with highlights
        self._apply_excel_highlights(output_path)

    def _apply_excel_highlights(self, output_path):
        wb = load_workbook(output_path)
        for sheet_name in wb.sheetnames:
            if sheet_name == "Audit_Summary":
                continue
            
            ws = wb[sheet_name]
            # detect duplicate rows
            data = list(ws.values)
            rows = data[1:]
            
            seen = set()
            duplicate_indices = set()
            for idx, row in enumerate(rows, start=2):
                if row in seen:
                    duplicate_indices.add(idx)
                else:
                    seen.add(row)
            
            for row in ws.iter_rows(min_row=2):
                row_idx = row[0].row
                if row_idx in duplicate_indices:
                    for cell in row:
                        cell.fill = YELLOW_FILL
                
                for cell in row:
                    if cell.value is None or str(cell.value).strip() == "":
                        cell.fill = BLUE_FILL
        
        # Reorder sheets
        new_order = ["Audit_Summary", "Original_Data"] + [
            s for s in wb.sheetnames if s not in ["Audit_Summary", "Original_Data"]
        ]
        wb._sheets = [wb[s] for s in new_order]
        wb.save(output_path)

    def generate_html(self, output_path):
        results = self.auditor.results
        summary = self.auditor.summary
        
        html_content = f"""
        <html>
        <head>
            <title>CMDB Audit Report - {datetime.now().strftime('%Y-%m-%d %H:%M')}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f4f4f9; }}
                h1, h2 {{ color: #333; }}
                table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; background-color: #fff; }}
                th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                th {{ background-color: #4CAF50; color: white; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                .score {{ font-weight: bold; font-size: 1.2em; }}
                .good {{ color: green; }}
                .fair {{ color: orange; }}
                .poor {{ color: red; }}
                .card {{ background: white; padding: 15px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); margin-bottom: 20px; }}
            </style>
        </head>
        <body>
            <h1>Data Audit Executive Summary</h1>
            <div class="card">
                <h2>Overview</h2>
                <table>
                    <tr>
                        <th>Category</th>
                        <th>Total Rows</th>
                        <th>Duplicates</th>
                        <th>Incomplete Rows</th>
                        <th>Quality Score</th>
                    </tr>
                    {"".join([f"<tr><td>{s['Category']}</td><td>{s['Total Rows']}</td><td>{s['Duplicates']}</td><td>{s['Incomplete Rows']}</td><td class='score {self._get_score_class(s['Quality Score'])}'>{s['Quality Score']}%</td></tr>" for s in summary])}
                </table>
            </div>
            
            <h2>Detailed Findings</h2>
        """
        
        for name, res in results.items():
            html_content += f"""
            <div class="card">
                <h3>{name}</h3>
                <p><b>Total Rows:</b> {res['total_rows']}</p>
                <p><b>Duplicates:</b> {res['duplicate_count']}</p>
                <p><b>Incomplete Rows:</b> {res['incomplete_rows']}</p>
                
                <h4>Missing Values by Column</h4>
                <ul>
                    {"".join([f"<li>{col}: {count}</li>" for col, count in res['missing_values'].items()]) if res['missing_values'] else "<li>None</li>"}
                </ul>
                
                <h4>Type Mismatches</h4>
                <ul>
                    {"".join([f"<li>{col}: {', '.join(types)}</li>" for col, types in res['type_mismatches'].items()]) if res['type_mismatches'] else "<li>None</li>"}
                </ul>

                <h4>Anomalies (Outliers)</h4>
                <ul>
                    {"".join([f"<li>{col}: {count} outliers detected</li>" for col, count in res['outliers'].items()]) if res['outliers'] else "<li>None</li>"}
                </ul>
            </div>
            """
        
        html_content += """
        </body>
        </html>
        """
        
        with open(output_path, "w") as f:
            f.write(html_content)

    def _get_score_class(self, score):
        if score >= 90: return "good"
        if score >= 70: return "fair"
        return "poor"
